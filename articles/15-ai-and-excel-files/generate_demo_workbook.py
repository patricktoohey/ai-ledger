"""
generate_demo_workbook.py

Generates a realistic multi-tab Excel workbook (monthly_financial_workbook.xlsx)
that mirrors a typical messy accounting file.

Tabs:
  - Raw_Data: ~60 rows of transaction-level data for Jan-Feb 2026
  - Pivot_Summary: Department-by-month totals (intentionally slightly wrong)
  - Dashboard: Placeholder headers and merged cells
  - Notes: Free-text notes from various team members

Usage:
    python generate_demo_workbook.py
"""

import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Seed for reproducibility
# ---------------------------------------------------------------------------
random.seed(42)

# ---------------------------------------------------------------------------
# Reference data
# ---------------------------------------------------------------------------
DEPARTMENTS = ["Finance", "Operations", "Marketing", "IT", "HR"]

# (Category, typical GL account prefix, amount low, amount high)
CATEGORY_POOL = [
    ("Rent", "6010", 8000, 25000),
    ("SaaS Subscriptions", "6210", 200, 4500),
    ("Travel", "6310", 150, 3800),
    ("Office Supplies", "6110", 150, 1200),
    ("Professional Services", "6410", 1500, 18000),
    ("Utilities", "6020", 350, 2800),
    ("Insurance", "6030", 1200, 6000),
    ("Training & Development", "6510", 250, 3500),
    ("Telecom", "6220", 180, 1600),
    ("Meals & Entertainment", "6320", 150, 950),
]

VENDORS = {
    "Rent": ["Metro Commercial Realty", "Beacon Property Group"],
    "SaaS Subscriptions": [
        "Salesforce", "Microsoft 365", "Slack Technologies",
        "Zoom Video Communications", "Adobe Systems", "QuickBooks Online",
        "Dropbox Business", "HubSpot",
    ],
    "Travel": [
        "Delta Air Lines", "United Airlines", "Marriott Hotels",
        "Hilton Worldwide", "Enterprise Rent-A-Car", "Uber Business",
    ],
    "Office Supplies": [
        "Staples", "Office Depot", "Amazon Business", "W.B. Mason",
    ],
    "Professional Services": [
        "Deloitte Consulting", "KPMG Advisory", "Baker McKenzie LLP",
        "Moss Adams LLP", "Plante Moran",
    ],
    "Utilities": [
        "Pacific Gas & Electric", "ConEdison", "Duke Energy",
    ],
    "Insurance": [
        "Hartford Financial", "Zurich Insurance", "Travelers Companies",
    ],
    "Training & Development": [
        "Coursera for Business", "LinkedIn Learning", "AICPA",
        "Becker Professional Education",
    ],
    "Telecom": [
        "AT&T Business", "Verizon Enterprise", "T-Mobile for Business",
    ],
    "Meals & Entertainment": [
        "DoorDash Corporate", "Grubhub", "Aramark Catering",
    ],
}

DESCRIPTIONS = {
    "Rent": ["Monthly office lease - {dept}", "Warehouse space rental"],
    "SaaS Subscriptions": [
        "Annual license renewal", "Monthly subscription",
        "Seat expansion - {dept}", "Platform upgrade",
    ],
    "Travel": [
        "Client site visit - {dept}", "Conference attendance",
        "Quarterly team offsite", "Airport transfer",
    ],
    "Office Supplies": [
        "Printer toner and paper", "Desk accessories",
        "Breakroom supplies", "Filing and storage",
    ],
    "Professional Services": [
        "Q1 advisory engagement", "Tax compliance review",
        "Audit support services", "Legal retainer - monthly",
    ],
    "Utilities": [
        "Electric - main office", "Water and sewer",
        "Natural gas - heating",
    ],
    "Insurance": [
        "General liability premium", "D&O policy renewal",
        "Workers comp quarterly",
    ],
    "Training & Development": [
        "CPE credits - staff", "Leadership workshop",
        "Online learning platform", "Certification prep course",
    ],
    "Telecom": [
        "Mobile plan - {dept}", "Office internet service",
        "Conference line subscription",
    ],
    "Meals & Entertainment": [
        "Team lunch - {dept}", "Client dinner",
        "All-hands catering",
    ],
}


def _random_date(month: int) -> str:
    """Return a random date string in M/D/YYYY format for Jan or Feb 2026."""
    max_day = 31 if month == 1 else 28
    day = random.randint(1, max_day)
    return f"{month}/{day}/2026"


def _generate_transactions(count: int = 60) -> list[dict]:
    """Build a list of realistic transaction dictionaries."""
    rows = []
    for _ in range(count):
        month = random.choice([1, 2])
        dept = random.choice(DEPARTMENTS)
        cat, gl_prefix, lo, hi = random.choice(CATEGORY_POOL)
        vendor = random.choice(VENDORS[cat])
        desc_template = random.choice(DESCRIPTIONS[cat])
        description = desc_template.format(dept=dept)
        amount = round(random.uniform(lo, hi), 2)
        gl_account = f"{gl_prefix}-{random.randint(100, 999)}"

        rows.append({
            "Date": _random_date(month),
            "Department": dept,
            "Category": cat,
            "Vendor": vendor,
            "Description": description,
            "Amount": amount,
            "GL_Account": gl_account,
        })

    # Sort by date string so the sheet looks chronologically ordered
    rows.sort(key=lambda r: (int(r["Date"].split("/")[0]),
                              int(r["Date"].split("/")[1])))
    return rows


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="002639", end_color="002639",
                          fill_type="solid")  # Deep Navy
TEAL_FILL = PatternFill(start_color="005F6F", end_color="005F6F",
                         fill_type="solid")  # Ocean Teal
LIGHT_FILL = PatternFill(start_color="E8F4F2", end_color="E8F4F2",
                          fill_type="solid")
CURRENCY_FMT = '#,##0.00'
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header_row(ws, col_count: int) -> None:
    """Apply formatting to row 1 as header."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _auto_width(ws, min_width: int = 10, max_width: int = 30) -> None:
    """Set column widths based on content length."""
    for col_cells in ws.columns:
        lengths = []
        for cell in col_cells:
            if cell.value is not None:
                lengths.append(len(str(cell.value)))
        if lengths:
            best = min(max(max(lengths) + 2, min_width), max_width)
            ws.column_dimensions[
                get_column_letter(col_cells[0].column)
            ].width = best


# ---------------------------------------------------------------------------
# Tab builders
# ---------------------------------------------------------------------------

def _build_raw_data(wb: Workbook, transactions: list[dict]) -> None:
    """Create the Raw_Data tab with transaction rows."""
    ws = wb.active
    ws.title = "Raw_Data"

    headers = ["Date", "Department", "Category", "Vendor",
               "Description", "Amount", "GL_Account"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for i, txn in enumerate(transactions, start=2):
        ws.cell(row=i, column=1, value=txn["Date"])
        ws.cell(row=i, column=2, value=txn["Department"])
        ws.cell(row=i, column=3, value=txn["Category"])
        ws.cell(row=i, column=4, value=txn["Vendor"])
        ws.cell(row=i, column=5, value=txn["Description"])
        amt_cell = ws.cell(row=i, column=6, value=txn["Amount"])
        amt_cell.number_format = CURRENCY_FMT
        ws.cell(row=i, column=7, value=txn["GL_Account"])

        # Alternate row shading
        if i % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = LIGHT_FILL

        # Thin borders on every cell
        for col in range(1, len(headers) + 1):
            ws.cell(row=i, column=col).border = THIN_BORDER

    # Freeze top row
    ws.freeze_panes = "A2"
    _auto_width(ws)


def _build_pivot_summary(wb: Workbook, transactions: list[dict]) -> None:
    """Create a Pivot_Summary tab with department-by-month totals.

    Intentionally introduces small discrepancies so the totals do not
    exactly match Raw_Data -- this demonstrates why AI should always
    recalculate from the source tab.
    """
    ws = wb.create_sheet("Pivot_Summary")

    # Compute real totals from raw data
    totals: dict[tuple[str, int], float] = {}
    for txn in transactions:
        dept = txn["Department"]
        month = int(txn["Date"].split("/")[0])
        key = (dept, month)
        totals[key] = totals.get(key, 0.0) + txn["Amount"]

    # Introduce deliberate errors: nudge Marketing-Jan and IT-Feb
    if ("Marketing", 1) in totals:
        totals[("Marketing", 1)] += 247.50   # overstate by $247.50
    if ("IT", 2) in totals:
        totals[("IT", 2)] -= 183.00          # understate by $183.00

    # Layout
    headers = ["Department", "Jan 2026", "Feb 2026", "Total"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    grand_jan = 0.0
    grand_feb = 0.0
    for row_idx, dept in enumerate(DEPARTMENTS, start=2):
        jan = round(totals.get((dept, 1), 0.0), 2)
        feb = round(totals.get((dept, 2), 0.0), 2)
        dept_total = round(jan + feb, 2)
        grand_jan += jan
        grand_feb += feb

        ws.cell(row=row_idx, column=1, value=dept).font = Font(bold=True)
        for col, val in [(2, jan), (3, feb), (4, dept_total)]:
            c = ws.cell(row=row_idx, column=col, value=val)
            c.number_format = CURRENCY_FMT
            c.border = THIN_BORDER

    # Grand total row
    gt_row = len(DEPARTMENTS) + 2
    ws.cell(row=gt_row, column=1, value="Grand Total").font = Font(
        bold=True, size=12)
    for col, val in [(2, round(grand_jan, 2)),
                     (3, round(grand_feb, 2)),
                     (4, round(grand_jan + grand_feb, 2))]:
        c = ws.cell(row=gt_row, column=col, value=val)
        c.number_format = CURRENCY_FMT
        c.font = Font(bold=True, size=12)
        c.border = Border(top=Side(style="double"),
                          bottom=Side(style="double"),
                          left=Side(style="thin"),
                          right=Side(style="thin"))

    _auto_width(ws, min_width=14)

    # Add a note about the discrepancy (hidden in plain sight)
    note_row = gt_row + 2
    note_cell = ws.cell(row=note_row, column=1,
                        value="Source: Raw_Data tab (manual pivot)")
    note_cell.font = Font(italic=True, color="888888", size=9)


def _build_dashboard(wb: Workbook) -> None:
    """Create a Dashboard tab with placeholder headers and merged cells."""
    ws = wb.create_sheet("Dashboard")

    # Title banner
    ws.merge_cells("A1:F2")
    title = ws["A1"]
    title.value = "Monthly Spend Overview"
    title.font = Font(name="Calibri", bold=True, size=18, color="FFFFFF")
    title.fill = PatternFill(start_color="002639", end_color="002639",
                             fill_type="solid")
    title.alignment = Alignment(horizontal="center", vertical="center")

    # Subtitle
    ws.merge_cells("A3:F3")
    sub = ws["A3"]
    sub.value = "January - February 2026"
    sub.font = Font(size=12, italic=True, color="005F6F")
    sub.alignment = Alignment(horizontal="center")

    # Section headers
    sections = [
        (5, "A", "D", "Department Breakdown"),
        (5, "E", "F", "Top Vendors"),
        (12, "A", "D", "Month-over-Month Trend"),
        (12, "E", "F", "Budget vs Actual"),
    ]
    for row, start_col, end_col, label in sections:
        ws.merge_cells(f"{start_col}{row}:{end_col}{row}")
        cell = ws[f"{start_col}{row}"]
        cell.value = label
        cell.font = Font(bold=True, size=13, color="FFFFFF")
        cell.fill = TEAL_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Placeholder text beneath each section
    placeholders = [
        (6, 1, "[Chart placeholder -- department pie chart]"),
        (6, 5, "[Chart placeholder -- top 10 vendors]"),
        (13, 1, "[Chart placeholder -- line chart by month]"),
        (13, 5, "[Chart placeholder -- variance bars]"),
    ]
    for row, col, text in placeholders:
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(italic=True, color="AAAAAA", size=10)

    # Set some column widths so it looks intentional
    for col_letter in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 18

    ws.sheet_properties.tabColor = "3ABFB9"


def _build_notes(wb: Workbook) -> None:
    """Create a Notes tab with free-text annotations."""
    ws = wb.create_sheet("Notes")

    notes = [
        ("A1", "Updated by Sarah 1/15/2026"),
        ("A3", "Q1 forecast pending -- waiting on ops budget from Mike"),
        ("A5", "Check Marketing variance with CFO"),
        ("A7", "IT hardware refresh moved to Q2 per board approval"),
        ("A9", "Rent increase effective March 2026 -- new lease signed 12/2025"),
        ("A11", "Professional services spike in Jan due to year-end audit fees"),
        ("A13", "Travel policy reminder: pre-approval required for trips > $2,000"),
    ]

    # Header
    ws.merge_cells("A1:D1")  # we won't use this merge, put header above notes
    ws.unmerge_cells("A1:D1")

    header_cell = ws.cell(row=1, column=1, value="Team Notes & Follow-ups")
    header_cell.font = Font(bold=True, size=14, color="002639")
    header_cell.alignment = Alignment(vertical="center")

    ws.column_dimensions["A"].width = 65

    for i, (_, text) in enumerate(notes):
        row = i + 3  # start on row 3 to leave space after header
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(size=11)
        if i % 2 == 0:
            c.fill = LIGHT_FILL

    ws.sheet_properties.tabColor = "FFD75E"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    transactions = _generate_transactions(count=60)

    wb = Workbook()

    _build_raw_data(wb, transactions)
    _build_pivot_summary(wb, transactions)
    _build_dashboard(wb)
    _build_notes(wb)

    output_path = Path(__file__).parent / "monthly_financial_workbook.xlsx"
    wb.save(str(output_path))
    print(f"Workbook saved to {output_path}")
    print(f"  Raw_Data rows:    {len(transactions)}")
    print(f"  Departments:      {', '.join(DEPARTMENTS)}")
    print(f"  Months covered:   Jan 2026, Feb 2026")
    print(f"  Pivot_Summary:    Intentional discrepancies in Marketing (Jan) and IT (Feb)")


if __name__ == "__main__":
    main()
